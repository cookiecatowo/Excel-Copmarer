import openpyxl
from openpyxl.styles import PatternFill
# path1 = "testC.xlsx"
# path2 = "testD.xlsx"

def excelTest(path1, path2):
    try:
        file1 = openpyxl.load_workbook(path1)
        file2 = openpyxl.load_workbook(path2)
    except:
        return("檔案類型不正確")
    #print(file1.sheetnames)
    output = openpyxl.Workbook()
    output_sheet = output.active
    output_row = 2

    new = PatternFill(fill_type='solid',fgColor='00FF00')
    change = PatternFill(fill_type='solid',fgColor='FFFF00')
    delete = PatternFill(fill_type='solid',fgColor='FF0000')

    output_sheet.cell(1, 1).value = "新增"
    output_sheet.cell(1, 1).fill = new
    output_sheet.cell(1, 2).value = "修改"
    output_sheet.cell(1, 2).fill = change
    output_sheet.cell(1, 3).value = "刪除"
    output_sheet.cell(1, 3).fill = delete


    #查找出檔案1.2中所有工作表名稱
    sheetNames = []
    for i in file1.sheetnames:
        if i not in sheetNames:
            sheetNames.append(i)
    for i in file2.sheetnames:
        if i not in sheetNames:
            sheetNames.append(i)
        #print (sheetNames)

    # 工作表名稱查找檔案1.2工作表
    for sheetName in sheetNames:
        is_sheet_exist = True
        try:
            sheet1 = file1.get_sheet_by_name(sheetName)
        except:
            is_sheet_exist = False
            print("檔案1中未有名為'"+sheetName+"'之工作表")
        try:
            sheet2 = file2.get_sheet_by_name(sheetName)
        except:
            is_sheet_exist = False
            print("檔案2中未有名為'"+sheetName+"'之工作表")

        # 比對同工作表中資料
        if is_sheet_exist:
            #將資料表1中資料存入字典
            output_sheet.cell(output_row, 1).value = str(sheetName)
            output_row += 1

            originRows = sheet1.max_row
            originColumns = sheet1.max_column
            originList = {}
            for row in range(1, originRows +1):
                id = ""
                info = []
                id = sheet1.cell(row, 1).value
                for column in range(2, originColumns +1):
                    info.append(sheet1.cell(row, column).value)
                originList[id] = info
            #print(originList)

            #比對資料表2是否有資料
            for row in range(1, sheet2.max_row +1):
                id = sheet2.cell(row, 1).value
                if originList.__contains__(id):
                    info = originList[id]
                    is_change = ""
                    for column in range(2, originColumns +1):
                        if info[column-2] != sheet2.cell(row, column).value:
                            is_change = column
                            sheet2.cell(row, column).fill = change
                    if is_change != "":
                        for column in range(1, originColumns +1):
                            output_sheet.cell(output_row,column).value = sheet2.cell(row,column).value
                            if column == is_change:
                                output_sheet.cell(output_row, column).fill = change
                        output_row +=1
                    originList[id] = "Searched"
                else:
                    for column in range(1, originColumns +1):
                        output_sheet.cell(output_row,column).value = sheet2.cell(row,column).value
                        output_sheet.cell(output_row, column).fill = new
                    output_row += 1

            #將刪除資料寫入Excel
            for item, content in originList.items():
                if content != "Searched":
                    output_row += 1
                    print(item,content)
                    output_sheet.cell(output_row,1).value = item
                    output_sheet.cell(output_row,1).fill = delete
                    for i in range (0,len(content)):
                        output_sheet.cell(output_row,i+2).value = content[i]
                        output_sheet.cell(output_row,i+2).fill = delete
            output_row += 2
    path2Name = path2.split(".")[0]
    output.save(path2Name + "_compared.xlsx")
    return path2Name + "_compared.xlsx"
